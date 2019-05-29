from openpyxl import load_workbook,Workbook
import datetime
import shutil
import os

class Excel_operation():

    def __init__(self,pathname):
        self.pathname = pathname

    #从指定目录中搜索并返回xlsx文件列表
    def get_xlsx_list(self):
        res = []
        pending_path = os.path.join(self.pathname,'pending')
        try:
            for dirpath, dirname, filenames in os.walk(pending_path):
                for filename in filenames:
                    if filename.endswith(".xlsx") and not filename.startswith('~$'):
                        res.append(os.path.join(dirpath, filename))
        except Exception as e:
            print(e)
        return res

    def solved_backup(self,path_file):
        path_solved = os.path.join(self.pathname,'solved')
        try:
            if not os.path.exists(path_solved):
                os.mkdir(path_solved)
            shutil.move(path_file,path_solved)
        except Exception as e:
            print(e)



    #按行读取内容
    def read_by_row(self,filepath,sheet_num,data_only=True):
        results = []
        try:
            wb = load_workbook(filepath,data_only=data_only)
            sheets = wb.sheetnames
            sheet = wb[sheets[sheet_num]]
            rows = sheet.rows
        except Exception as e:
            print(e)
            return 0

        #生成字段列表
        fields = [cell.value.replace(' ', '') for cell in next(rows)]
        results.append(fields + ['UpdatedTime'])

        #读取数据
        for row in rows:
            result = []
            for cell in row:
                result.append(cell.value)
            result.append(datetime.date.today())
            results.append(result)
        return results


    # #按行写入所有
    # def write_by_rows(self,filename,sheet_name,datas):
    #     wb = Workbook()
    #     index = 0
    #     wb.create_sheet(sheet_name, index=index)
    #     sheet = wb[sheet_name]
    #     for row in datas:
    #         sheet.append(row)
    #     wb.save(os.path.join(self.pathname,filename))






