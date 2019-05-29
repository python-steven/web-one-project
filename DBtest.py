# from app.login.models import BudgetCodeForm
# from django.db import connection
# from app import restful
from openpyxl import load_workbook,Workbook
# from AEMSLite.app.login import BudgetCodeForm
import time
import os
import psycopg2

conn = psycopg2.connect(
                database="aemslite",
                host="127.0.0.1",
                user="postgres",
                password = "1234qwer!@#$QWER",
                port = 5432,

            )

cur=conn.cursor()

partname = 'C'
sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where c.relname = '%s' and a.attrelid = c.oid and a.attnum>0 and a.attrelid is not null ''' % 'PartItem'
sql9 = 'select "PartItem"."SN","PartItem"."PartName","PartItem"."NGRate","PartItem"."ErrorCounts","UsedTimes" from "PartItem" where "Id"=63'

# sql2 = 'select  COUNT("PartName") from "PartItem" where "PartName" in (select distinct "PartName" from "PartItem")'
sql2 ='SELECT "PartName", COUNT("PartName") FROM "PartItem" where '
sql3 =sql1+'where name in("SN","PartName","NGRate","ErrorCounts","UsedTimes")'
#正常
sql4 = sql2 + '"NGRate" < \'' + '0.01' + '\'group by "PartName"'
#预警
sql5 = sql2 + '"NGRate" >= \'' + '0.01' + '\'' \
      + 'AND "NGRate" <= \'' + '2' + '\'group by "PartName"'
#超标
sql6 = sql2 + '"NGRate" > \'' + '2' + '\'group by "PartName"'

sql7 = 'select "PartName" from "PartItem" where "PartName" like \'%{0}%\''.format(partname)

sql10 = 'SELECT "SN","PartName","NGRate","ErrorCounts","UsedTimes","Configuration"."Min","Configuration"."Max" FROM "PartItem" INNER JOIN "Configuration" on "Configuration"."Id"=2 AND "PartItem"."Id" IN (63,64,65)'
sql11 = 'select "Max","Min" from "Configuration" where "Id"=2 '

# print(sql1)
a = 32,33,34
sql2 = 'select "BudgetCodeForm"."ExternalNumberType","Department","Remark","AttachmentPath","BillingType","BudgetCode","ExternalNumber"' \
               ',"ApplyDate","ExternalNumberEffectiveDate","Pic","ProductName","Model","PurchaseType","UnitPrice"' \
               ',"Quantity","Unit","Currency","Customer","TypeOfMachine","ProjectCode","ApplyReason" from "BudgetCodeForm" '
#所有的SN按ErrorCode分类统计出各ErrorCode类的SN数量所有的SN按ErrorCode分类统计出各ErrorCode类的SN数量
sl= 'SELECT "ErrorCode", COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' GROUP BY "ErrorCode"'
#所有的SN按品名分类统计出各品名类的SN数量
sl2 = 'SELECT "PartName", COUNT("SN") FROM "PartItemResult" GROUP BY "PartName"'
#所有的result 为Fail的SN按fail次数落在user设定fail次数范围区间统计出这些SN（会有重复的SN）的数量
sl3_1 = 'SELECT COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' and "UsedTimes">0 and "UsedTimes" <1000'
sl3_2 = 'SELECT COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' and "UsedTimes">=1000 and "UsedTimes" <=1500'
sl3_3 = 'SELECT COUNT("SN") FROM "PartItemResult" where "Result"= \'FAIL\' and "UsedTimes">=1500 and "UsedTimes" <=2000'
#所有的result 为Fail的SN按品名分类统计出的SN（过滤重复的SN）的数量
sl4 = 'SELECT "PartName", COUNT("SN") FROM (select distinct "SN","PartName","Result" from "PartItemResult") as foo  where "Result"= \'FAIL\' GROUP BY "PartName"'
sql = 'SELECT "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","PN" FROM "PartItem" where "SN" = Z17002D98'
sq = 'SELECT "SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","Configuration"."Max","Configuration"."Max" FROM' \
               ' "PartItem" INNER JOIN "Configuration" on "Configuration"."Type"= "mt_count","Configuration"."Type"=\'mt_date\''
def tes():
    # cur.execute(sl)
    # data = cur.fetchall()
    # print(data)
    #
    # cur.execute(sl2)
    # data = cur.fetchall()
    # print(data)
    #
    # cur.execute(sl3_1)
    # data = cur.fetchall()
    # cur.execute(sl3_2)
    # data_2 = cur.fetchall()
    # cur.execute(sl3_3)
    # data_3 = cur.fetchall()
    # data.append(data_2[0])
    # data.append(data_3[0])
    # print(data)

    cur.execute(sq)
    data = cur.fetchall()
    print(data)
tes()

# time_num = int(time.time())
# time_num = str(time_num)
# sheet_name = "预算表单1"
# filename = 'download' + time_num + '.xlsx'
#
# wb = Workbook()
# index = 0
# wb.create_sheet(sheet_name, index=index)
# sheet = wb[sheet_name]
# for row in data:
#     sheet.append(row)
# wb.save(os.path.join('/home/AEMSLite/AEMSLite/report/', filename))


# file_url = request.build_absolute_uri(settings.MEDIA_CHANGE_URL + filename)
# download_url = []
# download_url[0] = file_url
# return restful.ok(data=download_url)

# sql3 = sql2 +'"Id" IN '+str(a)
# print(type(a))
# print(sql3)
# 产生表头信息
# data =[attr[0] for attr in data]
# # print(data.index("SN"))
# data1 =[]
# data1.append(data[data.index("SN")])
# data1.append(data[data.index("PartName")])
# data1.append('Min')
# data1.append('Max')
# data1.append(data[data.index("NGRate")])
# data1.append(data[data.index("ErrorCounts")])
# data1.append(data[data.index("UsedTimes")])
# data1.append("status")
# # print(data,len(data))
# print(data1)
# # #产生表信息
# cur.execute(sql10)
# data2 = cur.fetchall()
# data2.insert(0,data1)
# print(data2, len(data2))
# for row in data2:
#     if row[]
# budget_user = BudgetCodeForm.objects.filter(Id__in=[15,16]).values("Signer").distinct().count()
# print()


def post(self, request):
    bud_id = request.POST.get('bud_id')
    bud_depart = request.POST.get('Department')
    bud_req = request.POST.get('Remark')
    bud_num_type = request.POST.get('bud_num_type')
    bud_num = request.POST.get('bud_num')
    bud_time = request.POST.get('bud_time')
    bud_principal = request.POST.get('bud_principal')
    bud_machine_name = request.POST.get('bud_machine_name')
    bud_machine_type = request.POST.get('bud_machine_type')
    bud_request_type = request.POST.get('bud_request_type')
    bud_price = request.POST.get('bud_price')
    bud_qty = request.POST.get('bud_qty')
    bud_qty_type = request.POST.get('bud_qty_type')
    # bud_total_price = request.POST['bud_total_price']
    bud_money_type = request.POST.get('bud_money_type')
    bud_customer = request.POST.get('bud_customer')
    bud_mach_type = request.POST.get('bud_mach_type')
    bud_project_code = request.POST.get('bud_project_code')
    #
    bud_user = request.POST.get('bud_user')
    bud_reason = request.POST.get('bud_reason')
    created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # upload_tool = request.POST.get('upload_tool')
    own_id = request.session.get('user_Id')
    # name_current = own_id.Name

    # url_num = genurlnum()
    time_num = int(time.time())
    time_num = str(time_num)
    file = request.FILES.get('upload_file')
    file_name = file.name
    file_sp_name = file_name.split('.')[0]
    file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]

    file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
    with open(file_path, 'wb') as f:
        for chunk in file.chunks():
            f.write(chunk)

    file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
    if bud_id != "":
        try:
            depart = Department.objects.get(Department=bud_depart)
            depart_id = depart.Id
            user = User.objects.get(Name=bud_user)
            user_id = user.Id
            cus = Customer.objects.get(Customer=bud_customer)
            cus_id = cus.Id
            pic_user = User.objects.get(Name=bud_principal)
            pic_user_id = pic_user.Id
            BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id,Department=bud_depart,Remark=bud_req
                                                            , Attachment=file_sp_name,ApplyDate=created_time
                                                            , ExternalNumberEffectiveDate=bud_time
                                                            , ExternalNumberType=bud_num_type,ExternalNumber=bud_num
                                                            , PicId=pic_user_id, Pic=bud_principal,ProductName=bud_machine_name
                                                            , Model=bud_machine_type,PurchaseType=bud_request_type
                                                            , UnitPrice=bud_price, Quantity=bud_qty,Unit=bud_qty_type
                                                            , Currency=bud_money_type, CustomerId=cus_id
                                                            , Customer=bud_customer,TypeOfMachine=bud_mach_type
                                                            , ProjectCode=bud_project_code,ApplyReason=bud_reason,SignerId=user_id
                                                            , Signer=bud_user, Status='Process',CreatedTime=created_time
                                                            , UpdatedTime=UpdatedTime, OwnerId=own_id,AttachmentPath=file_url
                                                            )
            # 邮件发送创建的表单给签核的人去签核表单信息
            subject = "Notification: AEMSLite Budget_code signing form information"
            email_1 = user.Email
            email_2 = pic_user.Email
            apply_er = User.objects.get(Id=own_id).Name
            content = """
<pre>
Dear """ + bud_user + """,
    """ + apply_er + """modify a maintain_monitor code eform waiting for your review/approval.
    You can click the below link to approve or reject. 
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    Thank you!
                                        

    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
            mail.sendmail([email_1, email_2], content, subject)
            return restful.ok(message='BudgetCodeForm modify success')
        except:
            return restful.params_error(message="input information exist error")
    else:
        try:
            depart = Department.objects.get(Department=bud_depart)
            depart_id = depart.Id
            user = User.objects.get(Name=bud_user)
            user_id = user.Id
            cus = Customer.objects.get(Customer=bud_customer)
            cus_id = cus.Id
            pic_user = User.objects.get(Name=bud_principal)
            pic_user_id = pic_user.Id
            BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                          , Attachment=file_sp_name, ApplyDate=created_time
                                          , ExternalNumberEffectiveDate=bud_time
                                          , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                          , PicId=pic_user_id, Pic=bud_principal,ProductName=bud_machine_name
                                          , Model=bud_machine_type, PurchaseType=bud_request_type
                                          , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                          , Currency=bud_money_type, CustomerId=cus_id
                                          , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                          , ProjectCode=bud_project_code, ApplyReason=bud_reason,SignerId=user_id
                                          , Signer=bud_user, Status='Process', CreatedTime=created_time
                                          , UpdatedTime=UpdatedTime, OwnerId=own_id,AttachmentPath=file_url
                                          )
            # 邮件发送创建的表单给签核的人去签核表单信息
            subject = "Notification: AEMSLite Budget_code signing form information"
            email_1 = user.Email
            email_2 = pic_user.Email
            apply_er = User.objects.get(Id=own_id).Name
            content = """
<pre>
Dear """ + bud_user + """,
    """ + apply_er + """modify a maintain_monitor code eform waiting for your review/approval.
    You can click the below link to approve or reject. 
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    Thank you!


    THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
            mail.sendmail([email_1, email_2], content, subject)
            # num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
            # num1 = []
            # num1.insert(0, num)
            return restful.ok(message='BudgetCodeForm create success')
        except:
            return restful.params_error(message="input information exist error")



# 保存提交的表单信息
def Budget_form_save(request):
    if request.method == "POST":
        bud_id = request.POST.get('bud_id')
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal')
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        bud_request_type = request.POST.get('bud_request_type')
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        # bud_total_price = request.POST['bud_total_price']
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_project_code = request.POST.get('bud_project_code')

        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        own_id = request.session.get('user_Id')

        time_num = int(time.time())
        time_num = str(time_num)

        file = request.FILES.get('upload_file')
        file_name = file.name
        file_sp_name = file_name.split('.')[0]
        file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]

        file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)
        f.close()

        file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        if bud_id != "":
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id
                BudgetCodeForm.objects.filter(Id=bud_depart).update(DepartmentId=depart_id,Department=bud_depart,Remark=bud_req
                                                                    , ApplyDate=created_time,Attachment=file_sp_name
                                                                    , ExternalNumberEffectiveDate=bud_time
                                                                    , ExternalNumberType=bud_num_type,ExternalNumber=bud_num
                                                                    , PicId=pic_user_id, Pic=bud_principal
                                                                    , ProductName=bud_machine_name,Model=bud_machine_type
                                                                    , PurchaseType=bud_request_type,UnitPrice=bud_price
                                                                    , Quantity=bud_qty, Unit=bud_qty_type
                                                                    , Currency=bud_money_type,CustomerId=cus_id
                                                                    , Customer=bud_customer,TypeOfMachine=bud_mach_type
                                                                    , ProjectCode=bud_project_code,ApplyReason=bud_reason
                                                                    , SignerId=user_id, Signer=bud_user,Status='Draft'
                                                                    , CreatedTime=created_time,UpdatedTime=UpdatedTime
                                                                    , OwnerId=own_id,AttachmentPath=file_url
                                                                    )
                # num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
                # num1 = []
                # num1.insert(0, num)
                return restful.ok(message="BudgetCodeForm modify success")
            except:
                return restful.params_error(message="input information exist error")
        else:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id
                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart,Remark=bud_req
                                              , ApplyDate=created_time, Attachment=file_sp_name
                                              , ExternalNumberEffectiveDate=bud_time
                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                              , PicId=pic_user_id, Pic=bud_principal
                                              , ProductName=bud_machine_name, Model=bud_machine_type
                                              , PurchaseType=bud_request_type, UnitPrice=bud_price
                                              , Quantity=bud_qty, Unit=bud_qty_type
                                              , Currency=bud_money_type, CustomerId=cus_id
                                              , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                              , ProjectCode=bud_project_code, ApplyReason=bud_reason
                                              , SignerId=user_id, Signer=bud_user, Status='Draft'
                                              , CreatedTime=created_time, UpdatedTime=UpdatedTime
                                              , OwnerId=own_id, AttachmentPath=file_url
                                              )
                return restful.ok(message="BudgetCodeForm create success")
            except:
                return restful.params_error(message="input information exist error")
#save form to db
# def Budget_form_save(request):
#     if request.method == "POST":
#         bud_id = request.POST.get('bud_id')
#         bud_depart = request.POST.get('Department')
#         bud_req = request.POST.get('Remark')
#         bud_num_type = request.POST.get('bud_num_type')
#         bud_num = request.POST.get('bud_num')
#         bud_time = request.POST.get('bud_time')
#         bud_principal = request.POST.get('bud_principal')
#         bud_machine_name = request.POST.get('bud_machine_name')
#         bud_machine_type = request.POST.get('bud_machine_type')
#         bud_request_type = request.POST.get('bud_request_type')
#         bud_price = request.POST.get('bud_price')
#         bud_qty = request.POST.get('bud_qty')
#         bud_qty_type = request.POST.get('bud_qty_type')
#         # bud_total_price = request.POST['bud_total_price']
#         bud_money_type = request.POST.get('bud_money_type')
#         bud_customer = request.POST.get('bud_customer')
#         bud_mach_type = request.POST.get('bud_mach_type')
#         bud_project_code = request.POST.get('bud_project_code')
#         #
#         bud_user = request.POST.get('bud_user')
#         bud_reason = request.POST.get('bud_reason')
#         created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         # upload_tool = request.POST.get('upload_tool')
#         own_id = request.session.get('user_Id')
#
#         # url_num = genurlnum()
#         time_num = int(time.time())
#         time_num =str(time_num)
#
#         file = request.FILES.get('upload_file')
#         file_name = file.name
#         file_sp_name = file_name.split('.')[0]
#         file_ven_name = file_sp_name +time_num +'.' + file_name.split('.')[1]
#
#         file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
#         with open(file_path, 'wb') as f:
#             for chunk in file.chunks():
#                 f.write(chunk)
#
#         file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
#         if bud_id != "":
#             try:
#                 depart = Department.objects.get(Department=bud_depart)
#                 depart_id = depart.Id
#                 try:
#                     user = User.objects.get(Name=bud_user)
#                     user_id = user.Id
#                     try:
#                         cus = Customer.objects.get(Customer=bud_customer)
#                         cus_id = cus.Id
#                         try:
#                             pic_user = User.objects.get(Name=bud_principal)
#                             pic_user_id = pic_user.Id
#                             if pic_user.IsActivated is True:
#                                 BudgetCodeForm.objects.filter(Id=bud_depart).update(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
#                                                               , ApplyDate=created_time, Attachment=file_sp_name
#                                                               , ExternalNumberEffectiveDate=bud_time
#                                                               , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
#                                                               , PicId=pic_user_id, Pic=bud_principal
#                                                               , ProductName=bud_machine_name,Model=bud_machine_type
#                                                               , PurchaseType=bud_request_type, UnitPrice=bud_price
#                                                               , Quantity=bud_qty, Unit=bud_qty_type
#                                                               , Currency=bud_money_type, CustomerId=cus_id
#                                                               , Customer=bud_customer, TypeOfMachine=bud_mach_type
#                                                               , ProjectCode=bud_project_code, ApplyReason=bud_reason
#                                                               , SignerId=user_id,Signer=bud_user, Status='Draft'
#                                                               , CreatedTime=created_time, UpdatedTime=UpdatedTime
#                                                               , OwnerId=own_id,AttachmentPath=file_url
#                                                               )
#                                 # num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
#                                 # num1 = []
#                                 # num1.insert(0, num)
#                                 return restful.ok(message="BudgetCodeForm modify success")
#                             else:
#                                 return restful.ok(message="PIC user had not activate")
#                         except:
#                             try:
#                                 pic_user_2 = User.objects.get(EmployeeId=bud_principal)
#                                 pic_user_2_id = pic_user_2.Id
#                                 if pic_user_2.IsActivated is True:
#                                     BudgetCodeForm.objects.filter(Id=bud_depart).update(DepartmentId=depart_id, Department=bud_depart
#                                                                   , Remark=bud_req,ApplyDate=created_time
#                                                                   , Attachment=file_sp_name
#                                                                   , ExternalNumberEffectiveDate=bud_time
#                                                                   , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
#                                                                   , PicId=pic_user_2_id, Pic=bud_principal
#                                                                   , ProductName=bud_machine_name
#                                                                   , Model=bud_machine_type, PurchaseType=bud_request_type
#                                                                   , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
#                                                                   , Currency=bud_money_type, CustomerId=cus_id
#                                                                   , Customer=bud_customer, TypeOfMachine=bud_mach_type
#                                                                   , ProjectCode=bud_project_code, ApplyReason=bud_reason
#                                                                   ,SignerId=user_id, Signer=bud_user, Status='Draft'
#                                                                   ,CreatedTime=created_time,UpdatedTime=UpdatedTime
#                                                                   ,OwnerId=own_id,AttachmentPath=file_url
#                                                                   )
#                                     # num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
#                                     # num1 = []
#                                     # num1.insert(0, num)
#                                     return restful.ok(message="BudgetCodeForm modify success")
#                                 else:
#                                     return restful.ok(message="PIC user had not activate")
#                             except:
#                                 return restful.params_error(message="PIC need username or employee")
#                     except:
#                         return restful.params_error(message="Customer no exist")
#                 except:
#                     return restful.params_error(message="user no exist")
#             except:
#                 return restful.params_error(message="department no exist")
#         else:
#             try:
#                 depart = Department.objects.get(Department=bud_depart)
#                 depart_id = depart.Id
#                 try:
#                     user = User.objects.get(Name=bud_user)
#                     user_id = user.Id
#                     try:
#                         cus = Customer.objects.get(Customer=bud_customer)
#                         cus_id = cus.Id
#                         try:
#                             pic_user = User.objects.get(Name=bud_principal)
#                             pic_user_id = pic_user.Id
#                             if pic_user.IsActivated is True:
#                                 BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
#                                                               , ApplyDate=created_time, Attachment=file_sp_name
#                                                               , ExternalNumberEffectiveDate=bud_time
#                                                               , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
#                                                               , PicId=pic_user_id, Pic=bud_principal
#                                                               , ProductName=bud_machine_name,Model=bud_machine_type
#                                                               , PurchaseType=bud_request_type, UnitPrice=bud_price
#                                                               , Quantity=bud_qty, Unit=bud_qty_type
#                                                               , Currency=bud_money_type, CustomerId=cus_id
#                                                               , Customer=bud_customer, TypeOfMachine=bud_mach_type
#                                                               , ProjectCode=bud_project_code, ApplyReason=bud_reason
#                                                               , SignerId=user_id,Signer=bud_user, Status='Draft'
#                                                               , CreatedTime=created_time, UpdatedTime=UpdatedTime
#                                                               , OwnerId=own_id,AttachmentPath=file_url
#                                                               )
#                                 return restful.ok(message="BudgetCodeForm create success")
#                             else:
#                                 return restful.ok(message="PIC user had not activate")
#                         except:
#                             try:
#                                 pic_user_2 = User.objects.get(EmployeeId=bud_principal)
#                                 pic_user_2_id = pic_user_2.Id
#                                 if pic_user_2.IsActivated is True:
#                                     BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart
#                                                                   , Remark=bud_req,ApplyDate=created_time
#                                                                   , Attachment=file_sp_name
#                                                                   , ExternalNumberEffectiveDate=bud_time
#                                                                   , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
#                                                                   , PicId=pic_user_2_id, Pic=bud_principal
#                                                                   , ProductName=bud_machine_name
#                                                                   , Model=bud_machine_type, PurchaseType=bud_request_type
#                                                                   , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
#                                                                   , Currency=bud_money_type, CustomerId=cus_id
#                                                                   , Customer=bud_customer, TypeOfMachine=bud_mach_type
#                                                                   , ProjectCode=bud_project_code, ApplyReason=bud_reason
#                                                                   ,SignerId=user_id, Signer=bud_user, Status='Draft'
#                                                                   ,CreatedTime=created_time,UpdatedTime=UpdatedTime
#                                                                   ,OwnerId=own_id,AttachmentPath=file_url
#                                                                   )
#                                     return restful.ok(message="BudgetCodeForm create success")
#                                 else:
#                                     return restful.ok(message="PIC user had not activate")
#                             except:
#                                 return restful.params_error(message="PIC need username or employee")
#                     except:
#                         return restful.params_error(message="Customer no exist")
#                 except:
#                     return restful.params_error(message="user no exist")
#             except:
#                 return restful.params_error(message="department no exist")


#megre order information get and Rendering to html
  # def post(self,request):
    #     bud_id = request.POST.get('bud_id')
    #     bud_depart = request.POST.get('Department')
    #     bud_req = request.POST.get('Remark')
    #     bud_num_type = request.POST.get('bud_num_type')
    #     bud_num = request.POST.get('bud_num')
    #     bud_time = request.POST.get('bud_time')
    #     bud_principal = request.POST.get('bud_principal')
    #     bud_machine_name = request.POST.get('bud_machine_name')
    #     bud_machine_type = request.POST.get('bud_machine_type')
    #     bud_request_type = request.POST.get('bud_request_type')
    #     bud_price = request.POST.get('bud_price')
    #     bud_qty = request.POST.get('bud_qty')
    #     bud_qty_type = request.POST.get('bud_qty_type')
    #     # bud_total_price = request.POST['bud_total_price']
    #     bud_money_type = request.POST.get('bud_money_type')
    #     bud_customer = request.POST.get('bud_customer')
    #     bud_mach_type = request.POST.get('bud_mach_type')
    #     bud_project_code = request.POST.get('bud_project_code')
    #     #
    #     bud_user = request.POST.get('bud_user')
    #     bud_reason = request.POST.get('bud_reason')
    #     created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #     # upload_tool = request.POST.get('upload_tool')
    #     own_id = request.session.get('user_Id')
    #     # name_current = own_id.Name
    #
    #     # url_num = genurlnum()
    #     time_num = int(time.time())
    #     time_num = str(time_num)
    #     file = request.FILES.get('upload_file')
    #     file_name = file.name
    #     file_sp_name =file_name.split('.')[0]
    #     file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]
    #
    #     file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
    #     with open(file_path, 'wb') as f:
    #         for chunk in file.chunks():
    #             f.write(chunk)
    #
    #     file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
    #     if bud_id != "":
    #         try:
    #             depart = Department.objects.get(Department=bud_depart)
    #             depart_id = depart.Id
    #             try:
    #                 user = User.objects.get(Name=bud_user)
    #                 user_id = user.Id
    #                 try:
    #                     cus = Customer.objects.get(Customer=bud_customer)
    #                     cus_id = cus.Id
    #                     try:
    #                         pic_user = User.objects.get(Name=bud_principal)
    #                         pic_user_id = pic_user.Id
    #                         if pic_user.IsActivated is True:
    #                             BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id, Department=bud_depart,
    #                                                           Remark=bud_req
    #                                                           , Attachment=file_sp_name, ApplyDate=created_time
    #                                                           , ExternalNumberEffectiveDate=bud_time
    #                                                           , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
    #                                                           , PicId=pic_user_id, Pic=bud_principal,
    #                                                           ProductName=bud_machine_name
    #                                                           , Model=bud_machine_type, PurchaseType=bud_request_type
    #                                                           , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
    #                                                           , Currency=bud_money_type, CustomerId=cus_id
    #                                                           , Customer=bud_customer, TypeOfMachine=bud_mach_type
    #                                                           , ProjectCode=bud_project_code, ApplyReason=bud_reason,
    #                                                           SignerId=user_id
    #                                                           , Signer=bud_user, Status='Process',
    #                                                           CreatedTime=created_time
    #                                                           , UpdatedTime=UpdatedTime, OwnerId=own_id,
    #                                                           AttachmentPath=file_url
    #                                                           )
    #                             # 邮件发送创建的表单给签核的人去签核表单信息
    #                             subject = "Notification: AEMSLite Budget_code signing form information"
    #                             email_1 = user.Email
    #                             email_2 = pic_user.Email
    #                             content = """
    #         <pre>
    #         Dear """ + bud_user + """,
    #             You have a Budget_Code Form application need you sign and please seen the below link address:
    #             Budget_code apply department:""" + bud_depart + """
    #             Budget_code apply charger:""" + bud_principal + """
    #             Budget_code apply product name:""" + bud_machine_name + """
    #             Please click the link below to signed the apply:
    #             <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    #
    #
    #             =====================================================================================================
    #             THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    #         </pre>
    #         """
    #                             mail.sendmail([email_1, email_2], content, subject)
    #                             return restful.ok(message='BudgetCodeForm modify success')
    #                         else:
    #                             return restful.ok(message="PIC user had not activate")
    #                     except:
    #                         try:
    #                             pic_user_2 = User.objects.get(EmployeeId=bud_principal)
    #                             pic_user_2_id = pic_user_2.Id
    #                             if pic_user_2.IsActivated is True:
    #                                 BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id, Department=bud_depart,
    #                                                               Remark=bud_req
    #                                                               , Attachment=file_sp_name, ApplyDate=created_time
    #                                                               , ExternalNumberEffectiveDate=bud_time
    #                                                               , ExternalNumberType=bud_num_type,
    #                                                               ExternalNumber=bud_num
    #                                                               , PicId=pic_user_2_id, Pic=bud_principal
    #                                                               , ProductName=bud_machine_name
    #                                                               , Model=bud_machine_type,
    #                                                               PurchaseType=bud_request_type
    #                                                               , UnitPrice=bud_price, Quantity=bud_qty,
    #                                                               Unit=bud_qty_type
    #                                                               , Currency=bud_money_type, CustomerId=cus_id
    #                                                               , Customer=bud_customer, TypeOfMachine=bud_mach_type
    #                                                               , ProjectCode=bud_project_code, ApplyReason=bud_reason
    #                                                               , SignerId=user_id, Signer=bud_user, Status='Process'
    #                                                               , CreatedTime=created_time
    #                                                               , UpdatedTime=UpdatedTime, OwnerId=own_id
    #                                                               , AttachmentPath=file_url
    #                                                               )
    #                                 # 邮件发送修改内容
    #                                 subject = "Notification: AEMSLite Budget_code signing form information"
    #                                 email_1 = user.Email
    #                                 email_2 = pic_user_2.Email
    #                                 content = """
    #         <pre>
    #         Dear """ + bud_user + """,
    #         You have a Budget_Code Form application need you sign and please seen the below link address:
    #         Budget_code apply department:""" + bud_depart + """
    #         Budget_code apply charger:""" + bud_principal + """
    #         Budget_code apply product name:""" + bud_machine_name + """
    #         Please click the link below to signed the apply:
    #         <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    #
    #
    #         =====================================================================================================
    #         THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    #         </pre>
    #         """
    #                                 mail.sendmail([email_1, email_2], content, subject)
    #                                 return restful.ok(message='BudgetCodeForm modify success')
    #                             else:
    #                                 return restful.ok(message="PIC user had not activate")
    #                         except:
    #                             return restful.params_error(message="PIC need username or employee")
    #                 except:
    #                     return restful.params_error(message="Customer no exist")
    #             except:
    #                 return restful.params_error(message="user no exist")
    #         except:
    #             return restful.params_error(message="department no exist")
    #     else:
    #         try:
    #             depart = Department.objects.get(Department=bud_depart)
    #             depart_id = depart.Id
    #             try:
    #                 user = User.objects.get(Name=bud_user)
    #                 user_id = user.Id
    #                 try:
    #                     cus = Customer.objects.get(Customer=bud_customer)
    #                     cus_id = cus.Id
    #                     try:
    #                         pic_user = User.objects.get(Name=bud_principal)
    #                         pic_user_id = pic_user.Id
    #                         if pic_user.IsActivated is True:
    #                             BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
    #                                                   , Attachment=file_sp_name, ApplyDate=created_time
    #                                                   , ExternalNumberEffectiveDate=bud_time
    #                                                   , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
    #                                                   , PicId=pic_user_id,Pic=bud_principal, ProductName=bud_machine_name
    #                                                   , Model=bud_machine_type,PurchaseType=bud_request_type
    #                                                   , UnitPrice=bud_price, Quantity=bud_qty,Unit=bud_qty_type
    #                                                   , Currency=bud_money_type, CustomerId=cus_id
    #                                                   , Customer=bud_customer, TypeOfMachine=bud_mach_type
    #                                                   , ProjectCode=bud_project_code, ApplyReason=bud_reason,SignerId=user_id
    #                                                   , Signer=bud_user, Status='Process', CreatedTime=created_time
    #                                                   , UpdatedTime=UpdatedTime, OwnerId=own_id,AttachmentPath=file_url
    #                                                   )
    # #邮件发送创建的表单给签核的人去签核表单信息
    #                             subject = "Notification: AEMSLite Budget_code signing form information"
    #                             email_1 = user.Email
    #                             email_2 = pic_user.Email
    #                             content = """
    # <pre>
    # Dear """ + bud_user + """,
    #     You have a Budget_Code Form application need you sign and please seen the below link address:
    #     Budget_code apply department:""" + bud_depart + """
    #     Budget_code apply charger:""" + bud_principal + """
    #     Budget_code apply product name:""" + bud_machine_name + """
    #     Please click the link below to signed the apply:
    #     <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    #
    #
    #     =====================================================================================================
    #     THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    # </pre>
    # """
    #                             mail.sendmail([email_1,email_2], content, subject)
    #                             num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
    #                             num1 = []
    #                             num1.insert(0, num)
    #                             return restful.ok(message='BudgetCodeForm create success',data=num1)
    #                         else:
    #                             return restful.ok(message="PIC user had not activate")
    #                     except:
    #                         try:
    #                             pic_user_2 = User.objects.get(EmployeeId=bud_principal)
    #                             pic_user_2_id = pic_user_2.Id
    #                             if pic_user_2.IsActivated is True:
    #                                 BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart,Remark=bud_req
    #                                                               , Attachment=file_sp_name, ApplyDate=created_time
    #                                                               , ExternalNumberEffectiveDate=bud_time
    #                                                               , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
    #                                                               , PicId=pic_user_2_id, Pic=bud_principal
    #                                                               , ProductName=bud_machine_name
    #                                                               , Model=bud_machine_type, PurchaseType=bud_request_type
    #                                                               , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
    #                                                               , Currency=bud_money_type, CustomerId=cus_id
    #                                                               , Customer=bud_customer, TypeOfMachine=bud_mach_type
    #                                                               , ProjectCode=bud_project_code, ApplyReason=bud_reason
    #                                                               , SignerId=user_id, Signer=bud_user, Status='Process'
    #                                                               , CreatedTime=created_time
    #                                                               , UpdatedTime=UpdatedTime, OwnerId=own_id
    #                                                               , AttachmentPath=file_url
    #                                                               )
    #                                 # 邮件发送修改内容
    #                                 subject = "Notification: AEMSLite Budget_code signing form information"
    #                                 email_1 = user.Email
    #                                 email_2 = pic_user_2.Email
    #                                 content = """
    # <pre>
    # Dear """ + bud_user + """,
    # You have a Budget_Code Form application need you sign and please seen the below link address:
    # Budget_code apply department:""" + bud_depart + """
    # Budget_code apply charger:""" + bud_principal + """
    # Budget_code apply product name:""" + bud_machine_name + """
    # Please click the link below to signed the apply:
    # <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    #
    #
    # =====================================================================================================
    # THIS EMAIL WAS SENT BY AEMSLite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    # </pre>
    # """
    #                                 mail.sendmail([email_1, email_2], content, subject)
    #                                 # num = BudgetCodeForm.objects.filter(Status="Process", PicId=user_id).count()
    #                                 # num1=[]
    #                                 # num1.insert(0,num)
    #                                 return restful.ok(message='BudgetCodeForm create success')
    #                             else:
    #                                 return restful.ok(message="PIC user had not activate")
    #                         except:
    #                             return restful.params_error(message="PIC need username or employee")
    #                 except:
    #                     return restful.params_error(message="Customer no exist")
    #             except:
    #                 return restful.params_error(message="user no exist")
    #         except:
    #             return restful.params_error(message="department no exist")