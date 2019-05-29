from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from datetime import datetime,timedelta,date
from django.conf import settings
import random
import string
import os
import time
from openpyxl import load_workbook,Workbook
import json
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#登入首页的时候需要验证信息
class IndexView(View):
    @csrf_exempt
    def get(self,request):
        try:
            id = request.session['user_Id']
            user = User.objects.get(Id=id)
            request.session.set_expiry(0)
            num = BudgetCodeForm.objects.filter(Status="Process",PicId=id,MergeId=None).count()\
                  + BudgetCodeForm.objects.exclude(MergeId=None).filter(Status="Process",PicId=id).distinct("MergeId").count()
            return render(request, "./index/main.html", {'user': user,'num':num})
        except:
            return HttpResponseRedirect("/login/")


#获取部门的信息和客户的信息
@csrf_exempt
def Budget_info_get(request):
    if request.method == "GET":
        data_dict={}
        cus_info = Customer.objects.exclude(IsActivated='False')
        cus_info = list(cus_info.values())
        depart_info = Department.objects.exclude(IsActivated='False')
        depart_info = list(depart_info.values())
        data_dict['customer'] =cus_info
        data_dict['department'] =depart_info
        return restful.ok(data=data_dict)

#检查用户和负责人的合法性
@csrf_exempt
def Budget_check_user(request):
    if request.method == "POST":
        check_user = request.POST['user_approve']
        try:
            user_c = User.objects.get(Name=check_user)
            if user_c:
                return restful.ok()
        except:
            return restful.params_error(message='the Signer not exist need admin check')

#检查负责人是否存在性
@csrf_exempt
def Budget_check_principal(request):
    if request.method == "POST":
        principal = request.POST['principal']
        try:
            user_c = User.objects.get(Name=principal)
            if user_c:
                return restful.ok()
        except:
            try:
                user_num = User.objects.get(EmployeeId=principal)
                if user_num:
                    return restful.ok()
            except:
                return restful.params_error(message='the PIC user not exist need admin check')


#表单信息的获取
class BudgetCodeApply(View):
    @csrf_exempt
    def get(self,request):
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(PicId=id).count()  # 总共多少条数据
            if number == "All":
                budgetcode_info = BudgetCodeForm.objects.filter(PicId=id).order_by("-UpdatedTime","MergeId") \
                    .values("Id", "BillingType", "Department", "ApplyDate", "Pic", "ProductName", "Signer", "Status",
                            "BudgetCode", "MergeId")
                budgetcode_info = list(budgetcode_info)
                dict_data['data'] = budgetcode_info
                dict_data['page_count'] = count
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budgetcode_info = BudgetCodeForm.objects.filter(PicId=id).order_by("-UpdatedTime","MergeId")\
                        .values("Id","BillingType","Department","ApplyDate","Pic","ProductName","Signer","Status","BudgetCode","MergeId")[(page-1)*number:number*page]
                    budgetcode_info = list(budgetcode_info)
                    dict_data['data'] = budgetcode_info
                    dict_data['page_count'] =page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages",data={'count':count})
        except Exception as e:
            return restful.params_error(message=repr(e))

    #提交信息送至签核人process
    @csrf_exempt
    def post(self, request):
        bud_id = request.POST.get('bud_id')
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal')
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        bud_request_type = request.POST.get('bud_request_type')
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        # bud_total_price = request.POST['bud_total_price']
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_project_code = request.POST.get('bud_project_code')
        #
        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # upload_tool = request.POST.get('upload_tool')
        own_id = request.session.get('user_Id')
        # name_current = own_id.Name
        # url_num = genurlnum()
        time_num = int(time.time())
        time_num = str(time_num)
        file = request.FILES.get('upload_file')
        if len(bud_time) == 0:
            bud_time =None
        if file:
            file_name = file.name
            file_sp_name = file_name.split('.')[0]
            file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]
            file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
            with open(file_path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            f.close()
            file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        else:
            file_sp_name=""
            file_url=""
        if bud_id != "":
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id
                BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id, Department=bud_depart,
                                                                Remark=bud_req
                                                                , Attachment=file_sp_name, ApplyDate=created_time
                                                                , ExternalNumberEffectiveDate=bud_time
                                                                , ExternalNumberType=bud_num_type,
                                                                ExternalNumber=bud_num
                                                                , PicId=pic_user_id, Pic=bud_principal,
                                                                ProductName=bud_machine_name
                                                                , Model=bud_machine_type, PurchaseType=bud_request_type
                                                                , UnitPrice=bud_price, Quantity=bud_qty,
                                                                Unit=bud_qty_type
                                                                , Currency=bud_money_type, CustomerId=cus_id
                                                                , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                                                , ProjectCode=bud_project_code, ApplyReason=bud_reason,
                                                                SignerId=user_id
                                                                , Signer=bud_user, Status='Process',
                                                                CreatedTime=created_time
                                                                , UpdatedTime=UpdatedTime, OwnerId=own_id,
                                                                AttachmentPath=file_url
                                                                )
                # 邮件发送创建的表单给签核的人去签核表单信息
                subject = "Apply Budge Code eForm to You Review"
                email_1 = user.Email
                email_2 = pic_user.Email
                apply_er = User.objects.get(Id=own_id).Name
                content = """
<pre>
Dear """ + bud_user + """,
    """ + apply_er + """modify a maintain_monitor code eform waiting for your review/approval.
    You can click the below link to approve or reject. Thank you!
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    


    THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System <a href="http://10.41.95.106:90/index/">
</pre>
"""
                mail.sendmail([email_1, email_2], content, subject)
                return restful.ok(message='BudgetCodeForm modify success')
            except Exception as e:
                return restful.params_error(message=repr(e))
        else:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id
                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                              , Attachment=file_sp_name, ApplyDate=created_time
                                              , ExternalNumberEffectiveDate=bud_time
                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                              , PicId=pic_user_id, Pic=bud_principal, ProductName=bud_machine_name
                                              , Model=bud_machine_type, PurchaseType=bud_request_type
                                              , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                              , Currency=bud_money_type, CustomerId=cus_id
                                              , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                              , ProjectCode=bud_project_code, ApplyReason=bud_reason, SignerId=user_id
                                              , Signer=bud_user, Status='Process', CreatedTime=created_time
                                              , UpdatedTime=UpdatedTime, OwnerId=own_id, AttachmentPath=file_url
                                              )
                # 邮件发送创建的表单给签核的人去签核表单信息
                subject = "Apply Budge Code eForm to You Review"
                email_1 = user.Email
                email_2 = pic_user.Email
                apply_er = User.objects.get(Id=own_id).Name
                content = """
<pre>
Dear """ + bud_user + """,
    """ + apply_er + """create a maintain_monitor code eform waiting for your review/approval.
    You can click the below link to approve or reject. Thank you!
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    


    THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System <a href="http://10.41.95.106:90/index/">
</pre>
"""
                mail.sendmail([email_1, email_2], content, subject)
                return restful.ok(message='BudgetCodeForm create success')
            except Exception as e:
                return restful.params_error(message=repr(e))

# 保存信息表单状态为Draft
@csrf_exempt
def Budget_form_save(request):
    if request.method == "POST":
        bud_id = request.POST.get('bud_id')
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal')
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        bud_request_type = request.POST.get('bud_request_type')
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        # bud_total_price = request.POST['bud_total_price']
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_project_code = request.POST.get('bud_project_code')

        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        own_id = request.session.get('user_Id')

        time_num = int(time.time())
        time_num = str(time_num)
        if len(bud_time) ==0:
            bud_time=None
        file = request.FILES.get('upload_file')
        if file:
            file_name = file.name
            file_sp_name = file_name.split('.')[0]
            file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]
            file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
            with open(file_path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            f.close()
            file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        else:
            file_sp_name=""
            file_url=""
        if len(bud_id) !=0:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id
                BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id,Department=bud_depart,Remark=bud_req
                                                                    , ApplyDate=created_time,Attachment=file_sp_name
                                                                    , ExternalNumberEffectiveDate=bud_time
                                                                    , ExternalNumberType=bud_num_type,ExternalNumber=bud_num
                                                                    , PicId=pic_user_id, Pic=bud_principal
                                                                    , ProductName=bud_machine_name,Model=bud_machine_type
                                                                    , PurchaseType=bud_request_type,UnitPrice=bud_price
                                                                    , Quantity=bud_qty, Unit=bud_qty_type
                                                                    , Currency=bud_money_type,CustomerId=cus_id
                                                                    , Customer=bud_customer,TypeOfMachine=bud_mach_type
                                                                    , ProjectCode=bud_project_code,ApplyReason=bud_reason
                                                                    , SignerId=user_id, Signer=bud_user,Status='Draft'
                                                                    , CreatedTime=created_time,UpdatedTime=UpdatedTime
                                                                    , OwnerId=own_id,AttachmentPath=file_url
                                                                    )
                return restful.ok(message="BudgetCodeForm modify success")
            except Exception as e:
                return restful.params_error(message=repr(e))
        else:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(Name=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Name=bud_principal)
                pic_user_id = pic_user.Id

                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart,Remark=bud_req
                                              , ApplyDate=created_time, Attachment=file_sp_name
                                              , ExternalNumberEffectiveDate=bud_time
                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                              , PicId=pic_user_id, Pic=bud_principal
                                              , ProductName=bud_machine_name, Model=bud_machine_type
                                              , PurchaseType=bud_request_type, UnitPrice=bud_price
                                              , Quantity=bud_qty, Unit=bud_qty_type
                                              , Currency=bud_money_type, CustomerId=cus_id
                                              , Customer=bud_customer, TypeOfMachine=bud_mach_type
                                              , ProjectCode=bud_project_code, ApplyReason=bud_reason
                                              , SignerId=user_id, Signer=bud_user, Status='Draft'
                                              , CreatedTime=created_time, UpdatedTime=UpdatedTime
                                              , OwnerId=own_id, AttachmentPath=file_url
                                              )
                return restful.ok(message="BudgetCodeForm create success")
            except Exception as e:
                return restful.params_error(message=repr(e))
#合并表单为Draft的信息获取
@csrf_exempt
def Budget_merge_order(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budgetcode_megre_info = BudgetCodeForm.objects.filter(Status='Draft',OwnerId=id).order_by("-UpdatedTime")\
                .values("Id","Department", "ApplyDate", "Pic","ProductName","Signer", "Status")
            budgetcode_megre_info =list(budgetcode_megre_info)
            return restful.ok(data=budgetcode_megre_info)
        except Exception as e:
            return restful.params_error(message=repr(e))

#合并开单的信息送至签核
@csrf_exempt
def merge_form_sub(request):
    if request.method == "POST":
        checked_id_array = request.POST.getlist('ids[]')
        own_id = request.session.get('user_Id')
        time_id =int(time.time())
        try:
            budget_user = BudgetCodeForm.objects.filter(Id__in=checked_id_array).values("Signer").distinct("Signer")
            singer_num = budget_user.count()
            if singer_num == 1:
                BudgetCodeForm.objects.filter(Id__in=list(checked_id_array)).update(MergeId=time_id, BillingType=1,Status="Process")
                # 邮件发送合并表单的要签核的信息给签核人
                subject = "Apply Budge Code eForm to You Review"
                budget_user =list(budget_user)
                sign_user = User.objects.get(Name=budget_user[0]['Signer'])
                email_1 = sign_user.Email
                sign_user_name = sign_user.Name
                apply_user = User.objects.get(Id=own_id)
                apply_user = apply_user.Name
                content = """
<pre>
Dear """ + sign_user_name + """,
    """+ apply_user +"""apply a maintain_monitor code eform waiting for your review/approval.
    You can click the below link to approve or reject. Thank you!
    <a href="http://10.41.95.106:90/index">index-sign AEMSLite</a>
    


    THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System <a href="http://10.41.95.106:90/index/">
</pre>
"""
                mail.sendmail([email_1,], content, subject)
                return restful.ok(message='BudgetCodeForm merged success')
            else:
                return restful.params_error(message="merged form signer different")
        except Exception as e:
            return restful.params_error(message=repr(e))
#修改之前先获取表里面的数据并且填入输入框中
@csrf_exempt
def budget_detail_modify(request):
    if request.method == "POST":
        detail_id = request.POST.get('id')
        try:
            detail_obj = BudgetCodeForm.objects.filter(Id=detail_id).values("ExternalNumberEffectiveDate","ExternalNumber")
            detail_obj=list(detail_obj)
            return restful.ok(data=detail_obj)
        except Exception as e:
            return restful.params_error(message=repr(e))


#修改表单信息
@csrf_exempt
def budget_modify_type(request):
    if request.method == "POST":
        modify_id = int(request.POST['modify_id'])
        modify_date = request.POST['modify_date']
        modify_number = request.POST['modify_number']
        try:
            modify_budget_ob = BudgetCodeForm.objects.get(Id=modify_id)
            mer_id =modify_budget_ob.MergeId
            if mer_id != None:
                BudgetCodeForm.objects.filter(MergeId=mer_id).update(ExternalNumberEffectiveDate=modify_date,ExternalNumber=modify_number)
                return restful.ok(message="modify form success")
            else:
                obj = BudgetCodeForm.objects.get(Id=modify_id)
                obj.ExternalNumberEffectiveDate = modify_date
                obj.ExternalNumber = modify_number
                obj.save()
                return restful.ok(message="modify form success")
        except Exception as e:
            return restful.params_error(message=repr(e))

#修改reject和draft的表单的获取信息
@csrf_exempt
def budget_modify_unique(request):
    if request.method == "POST":
        modify_unique_id = request.POST['modify_unique_id']
        try:
            modify_unique_ob = BudgetCodeForm.objects.filter(Id=modify_unique_id)
            modify_unique_ob = list(modify_unique_ob.values())
            return restful.ok(data=modify_unique_ob)
        except Exception as e:
            return restful.params_error(message=repr(e))

#复制表单信息
@csrf_exempt
def budget_copy_type(request):
    if request.method == "POST":
        copy_id = request.POST['copy_id']
        try:
            cop_ob = BudgetCodeForm.objects.get(Id=copy_id)
            BudgetCodeForm.objects.create(DepartmentId=cop_ob.DepartmentId, Department=cop_ob.Department
                                          , Remark=cop_ob.Remark, ApplyDate=cop_ob.ApplyDate
                                          , Attachment=cop_ob.Attachment
                                          , ExternalNumberEffectiveDate=cop_ob.ExternalNumberEffectiveDate
                                          , ExternalNumberType=cop_ob.ExternalNumberType, ExternalNumber=cop_ob.ExternalNumber
                                          , PicId=cop_ob.PicId, Pic=cop_ob.Pic
                                          , ProductName=cop_ob.ProductName
                                          , Model=cop_ob.Model, PurchaseType=cop_ob.PurchaseType
                                          , UnitPrice=cop_ob.UnitPrice, Quantity=cop_ob.Quantity, Unit=cop_ob.Unit
                                          , Currency=cop_ob.Currency, CustomerId=cop_ob.CustomerId
                                          , Customer=cop_ob.Customer, TypeOfMachine=cop_ob.TypeOfMachine
                                          , ProjectCode=cop_ob.ProjectCode, ApplyReason=cop_ob.ApplyReason
                                          , SignerId=cop_ob.SignerId, Signer=cop_ob.Signer, Status='Draft'
                                          , CreatedTime=cop_ob.CreatedTime, UpdatedTime=UpdatedTime
                                          , OwnerId=cop_ob.OwnerId, AttachmentPath=cop_ob.AttachmentPath
                                          )
            return restful.ok(message='copy form success')
        except Exception as e:
            return restful.params_error(message=repr(e))

#删除(取消)表单信息
@csrf_exempt
def budget_delete_type(request):
    if request.method == "POST":
        try:
            del_id = request.POST['del_id']
            budget_ob = BudgetCodeForm.objects.get(Id=del_id)
            if budget_ob.Status == "Draft" or budget_ob.Status == "Reject" or budget_ob.Status == "Cancel":
                BudgetCodeForm.objects.filter(Id=del_id).delete()
                return restful.ok(message="form delete success")

            if budget_ob.Status == "Process":
                BudgetCodeForm.objects.filter(Id=del_id).update(Status="Cancel",MergeId=None)
                cancel_budget = BudgetCodeForm.objects.get(Id=del_id)
                cancel_budget_Name =cancel_budget.Signer
                cancel_budget_principal= cancel_budget.Pic

                mail_user = User.objects.get(Name=cancel_budget_Name)
                mail_principal = User.objects.get(Name=cancel_budget_principal)
                email_1 = mail_user.Email
                email_2 = mail_principal.Email
                subject = "Reject Budge Code eForm"
                content = """
<pre>
Dear """ + str(cancel_budget_principal) + """,
    """+str(cancel_budget_Name)+"""has canceled the maintain_monitor code eform.
    The reason is:"""+str(cancel_budget.SignRemarks)+"""




    -------------------------------------------------------------------------------
    THIS EMAIL WAS SENT BY PTS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System <a href="http://10.41.95.106:90/index/">
</pre>
 """
                mail.sendmail([email_1, email_2], content, subject)
                return restful.ok(message='BudgetCodeForm cancel success',data={})
        except Exception as e:
            return restful.params_error(message=repr(e))

#签核单号的信息的获取函数
@csrf_exempt
def budget_singing_info(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budget_singing_data = BudgetCodeForm.objects.filter(Status='Process',SignerId=id).order_by("-UpdatedTime")\
                .values("Id","BillingType", "Department", "ApplyDate","Pic","ProductName", "Signer", "Status","MergeId")
            budget_singing_data = list(budget_singing_data)
            return restful.ok(data=budget_singing_data)
        except Exception as e:
            return restful.params_error(message=repr(e))

#签核表单内容
@csrf_exempt
def merge_signed(request):
    if request.method == "POST":
        bud_id = request.POST['bud_id']
        bud_merged_id = request.POST['bud_merged_id']
        bud_budgetcode = request.POST['budget_cod_text']
        bud_signremarks = request.POST['budget_text']
        sign_id = request.session['user_Id']
        try:
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)

            if user.Name != sign_budget_ob.Signer:
                return restful.params_error(message="you need connect admin")
            else:
                if bud_merged_id == "null":
                    bud_obj = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_obj.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode,
                                                                        SignRemarks=bud_signremarks, Status='Approve')
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        email_1 = mail_user.Email
                        subject = "Approve Budge Code eForm"
                        content = """
<pre>
Dear """ + sign_pincipal + """,
    """+sign_budget_ob.Signer+"""has approved your maintain_monitor code eform. 
     please seen the below link address:

    Please click the link below to see the detail information:
    <a href="http://10.41.95.106:90/index">index AEMS Lite</a>


-----------------------------------------------------------------------------------------
THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
 """
                        mail.sendmail([email_1, ], content, subject)
                        return restful.ok(message='BudgetCodeForm approved success')
                else:
                    bud_ob_merged = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_ob_merged.Status == "Cancel":
                        return restful.params_error("this form had Cancel")
                    else:
                        BudgetCodeForm.objects.filter(MergeId=bud_merged_id).update(BudgetCode=bud_budgetcode,
                                                                                    SignRemarks=bud_signremarks,
                                                                                    Status='Approve')
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        email_1 = mail_user.Email
                        subject = "Approve Budge Code eForm"
                        content = """
<pre>
Dear """ + sign_pincipal + """,
    """+sign_budget_ob.Signer+"""has approved your maintain_monitor code eform. 
     please seen the below link address:

    Please click the link below to see the detail information:
    <a href="http://10.41.95.106:90/index">index AEMS Lite</a>


-----------------------------------------------------------------------------------------
THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                        mail.sendmail([email_1, ], content, subject)
                        return restful.ok(message='BudgetCodeForm signed success')
        except Exception as e:
            return restful.params_error(message=repr(e))

#拒绝签核表单信息内容
@csrf_exempt
def merge_rejected(request):
    if request.method == "POST":
        bud_id = request.POST['bud_id']
        bud_merged_id = request.POST['bud_merged_id']
        bud_budgetcode = request.POST['budget_cod_text']
        bud_signremarks = request.POST['budget_text']
        sign_id = request.session['user_Id']
        try:
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)
            if user.Name != sign_budget_ob.Signer:
                return restful.params_error(message="you need connect admin")
            else:
                if bud_merged_id == "null":
                    if sign_budget_ob.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode,
                                                                        SignRemarks=bud_signremarks, Status='Reject')
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        email_1 = mail_user.Email
                        subject = "Reject Budge Code eForm"
                        content = """
<pre>
Dear """ + sign_pincipal + """,
    """+sign_budget_ob.Signer+"""has rejected your maintain_monitor code eform.
The reason is:"""+bud_signremarks+"""




-------------------------------------------------------------------------------
THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                        mail.sendmail([email_1, ], content, subject)
                        return restful.ok(message='BudgetCodeForm rejected')
                else:
                    if sign_budget_ob.Status == "Cancel":
                        return restful.params_error("this form had Cancel")
                    else:
                        BudgetCodeForm.objects.filter(MergeId=bud_merged_id).update(BudgetCode=bud_budgetcode,
                                                                                    SignRemarks=bud_signremarks,
                                                                                    Status='Reject')
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(Name=sign_pincipal)
                        # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Reject Budge Code eForm"
                        content = """
<pre>
Dear """ + sign_pincipal + """,
    """+sign_budget_ob.Signer+"""has rejected your maintain_monitor code eform.
The reason is:"""+bud_signremarks+"""




-------------------------------------------------------------------------------
THIS EMAIL WAS SENT BY AEMS Lite SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
</pre>
"""
                        mail.sendmail([email_1, ], content, subject)
                        # num = BudgetCodeForm.objects.filter(Status="Process", SignerId=sign_id).count()
                        # num1 = []
                        # num1.insert(0, num)
                        return restful.ok(message='BudgetCodeForm rejected')
        except Exception as e:
            return restful.params_error(message=repr(e))

#获取预算编码的内容get
@csrf_exempt
def merge_signed_finished(request):
    if request.method == "GET":
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject'),SignerId=id).count()
            if number== "All":
                budget_singed_info = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject'),SignerId=id).order_by("-UpdatedTime","MergeId")
                budget_singed_info = budget_singed_info.values("Id","BillingType", "Department", "ApplyDate"
                                                               , "Pic","ProductName", "Signer", "Status"
                                                               , "BudgetCode", "MergeId")
                budget_singed_info = list(budget_singed_info)
                dict_data['data']=budget_singed_info
                dict_data['page_count']=count
                return restful.ok(data=dict_data)
            if number !="All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budget_singed_info = BudgetCodeForm.objects.filter(Status__in=('Approve', 'Reject'),
                                                                       SignerId=id).order_by("-UpdatedTime", "MergeId")
                    budget_singed_info = budget_singed_info.values("Id", "BillingType", "Department", "ApplyDate"
                                                                   , "Pic", "ProductName", "Signer", "Status"
                                                                   , "BudgetCode", "MergeId")[(page-1)*number:number*page]
                    budget_singed_info = list(budget_singed_info)
                    budget_singed_info = list(budget_singed_info)
                    dict_data['data'] = budget_singed_info
                    dict_data['page_count'] = page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages",data={'count':count})
        except Exception as e:
            return restful.params_error(message=repr(e))

#获取预算编码信息并且生成报表的信息
@csrf_exempt
def merge_statement_detail(request):
    if request.method == "GET":
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(Status='Approve',SignerId=id).count()
            if number == "All":
                budget_statement_detail = BudgetCodeForm.objects.filter(Status='Approve',SignerId=id)
                budget_statement_detail = budget_statement_detail.values("Id","BillingType", "Department", "ApplyDate"
                                                                        ,"Pic","ProductName", "Signer", "Status"
                                                                        ,"BudgetCode", "MergeId")
                budget_statement_detail = list(budget_statement_detail)
                dict_data['data'] = budget_statement_detail
                dict_data['page_count'] = count
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budget_statement_detail = BudgetCodeForm.objects.filter(Status='Approve', SignerId=id)
                    budget_statement_detail = budget_statement_detail.values("Id", "BillingType", "Department","ApplyDate"
                                                                             , "Pic", "ProductName", "Signer", "Status"
                                                                             , "BudgetCode", "MergeId")[(page-1)*number:number*page]
                    budget_statement_detail = list(budget_statement_detail)
                    dict_data['data'] = budget_statement_detail
                    dict_data['page_count'] = page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages")
        except Exception as e:
            return restful.params_error(message=repr(e))

#筛选信息的获取
@csrf_exempt
def statement_query(request):
    if request.method == "POST":
        page = int(request.POST.get('page'))
        number = request.POST.get('num')
        query_billing_type = request.POST['query_billing_type']
        query_department = request.POST['query_department']
        query_start_date = request.POST['query_start_date']
        query_end_date = request.POST['query_end_date']
        query_pic = request.POST['query_pic']
        query_product_name = request.POST['query_product_name']
        query_signer = request.POST['query_signer']
        id = request.session['user_Id']
        dict_data={}
        try:
            sql ='select "Id","BillingType","Department","ApplyDate","Pic","ProductName","Signer","Status","BudgetCode"' \
                 ',"MergeId" FROM "BudgetCodeForm" WHERE "Status"=\'Approve\' AND "OwnerId"=\''+str(id)+'\' ' #查询状态为approve的对象
            sql2 = 'select count(*) FROM "BudgetCodeForm" WHERE "Status"=\'Approve\' AND "OwnerId"=\''+str(id)+'\' '
            cur = connection.cursor()
            cur.execute(sql2)
            count = cur.fetchall() #数量的总数
            if query_billing_type != "":
                sql = sql + 'AND "BillingType" = \''+query_billing_type +'\''
            if query_department != "":
                sql = sql+'and "Department" ilike \'%'+ query_department +'%\''
            if query_pic != "":
                sql = sql + 'AND "Pic" ilike \'%'+query_pic +'%\''
            if query_product_name != "":
                sql = sql + 'AND "ProductName" ilike \'%'+query_product_name +'%\''
            if query_signer != "":
                sql = sql + 'AND "Signer" ilike \'%'+query_signer +'%\''
            if query_start_date != "":
                sql = sql + 'and "ApplyDate" >=\'' + query_start_date + '\''
            if query_end_date != "":
                sql = sql + 'and "ApplyDate" <=\'' + query_end_date + '\''
            if number == 'All':
                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()
                dict_data['data'] = data
                dict_data['page_count'] = count[0][0]
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                count_page = count[0][0] // number  # 总数除以一页显示多少条，得到总的页数
                if count[0][0] % number > 0:
                    count_page += 1
                if page <= count_page:
                    sql = sql +' order by "ApplyDate" limit '+str(number)+' offset ' + str(page-1)
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data']=data
                    dict_data['page_count']=count_page
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages")
        except Exception as e:
            return restful.params_error(message=repr(e))

#详细信息的展示和加载
@csrf_exempt
def budget_code_detail(request):
    if request.method == "POST":
        detail_id = request.POST['detail_id']
        detail_merged_id = request.POST['detail_merged_id']
        detail_merged_id =str(detail_merged_id)
        try:
            try:
                detail_data = BudgetCodeForm.objects.filter(MergeId=detail_merged_id).values()
                detail_data = list(detail_data)
                return restful.ok(data=detail_data)
            except:
                detail_data = BudgetCodeForm.objects.filter(Id=detail_id).values()
                detail_data = list(detail_data)
                return restful.ok(data=detail_data)
        except:
            return restful.params_error(message='data fail ')

#生成报表文件的处理方式
@csrf_exempt
def statement_bring_info(request):
    if request.method == "POST":
        statement_ids = request.POST.getlist('statement_ids[]')
        # sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where
        #                                             c.relname = '%s' and a.attrelid = c.oid and a.attnum>0''' % 'BudgetCodeForm'
        #表头信息
        tb1 =['部门','備註/新增或損耗','鏈接','開單狀況','預算編號','PMCS单号','申請日期','PMCS簽核日期','PIC'
            ,'設備名稱/治具類型','規格/型號/版本','類別','單價','申請數量','單位','總費用(RMB)','實際發生金額(KRMB)'
            ,'RMB/USD','客戶','機種','ProjectCode','申請原因/用途']
        tb2 = ['部门', '備註/新增或損耗', '鏈接', '開單狀況', '預算編號', '201单号', '申請日期', '201簽核日期', 'PIC'
            , '設備名稱/治具類型', '規格/型號/版本', '類別', '單價', '申請數量', '單位', '總費用(RMB)', '實際發生金額(KRMB)'
            , 'RMB/USD', '客戶', '機種', 'ProjectCode', '申請原因/用途']
        sql2 = 'select "Department","Remark","AttachmentPath","BillingType","BudgetCode","ExternalNumberType"' \
               ',"ExternalNumber","ApplyDate","ExternalNumberEffectiveDate","Pic","ProductName","Model"' \
               ',"PurchaseType","UnitPrice","Quantity","Unit","Currency","Customer","TypeOfMachine"' \
               ',"ProjectCode","ApplyReason" from "BudgetCodeForm" where'
        if len(statement_ids) == 1:
            statement_ids= statement_ids[0]
            sql3 = sql2 + '"Id"=' + statement_ids

        else:
            statement_ids = tuple(statement_ids)
            statement_ids =str(statement_ids)
            sql3 = sql2 + '"Id" in ' + statement_ids

        cur = connection.cursor()
        cur.execute(sql3)
        data = list(cur.fetchall())
        data1=[]
        data2=[]
        for i in data:
            i=list(i)
            if i[7]:
                i[7]=str(i[7]).split(' ')[0]
            if i[8]:
                i[8]=str(i[8]).split(' ')[0]
            if i[5] == '1':
                if i[3] == '0':
                    i[3] = "單獨開單"
                if i[3] == '1':
                    i[3] = "合併開單"
                i.remove('1')
                sum = round(i[12] * i[13],2)
                i.insert(15,str(sum))
                if i[13] == "杂购":
                    i.insert(16,str('%.3f' %(sum/1000)))
                elif sum>600000:
                    i.insert(16,str('%.3f' %(sum/36000)))
                else:
                    i.insert(16,str('%.3f' %(sum/12000)))
                data1.append(i)
            if i[5] == '2':
                if i[3] == '0':
                    i[3] = "單獨開單"
                if i[3] == '1':
                    i[3] = "合併開單"
                i.remove('2')
                sum = round(i[12] * i[13],2)
                i.insert(15, str(sum))
                if i[13] == "杂购":
                    i.insert(16, str('%.3f' %(sum / 1000)))
                elif sum > 600000:
                    i.insert(16, str('%.3f' %(sum / 36000)))
                else:
                    i.insert(16, str('%.3f' %(sum / 12000)))
                data2.append(i)

        data1.insert(0,tb1)
        data2.insert(0,tb2)

        #写入文件
        time_num = int(time.time())
        time_num = str(time_num)
        filename = 'budget_code_report_' + time_num + '.xlsx'
        wb = Workbook()
        index = 0
        sheet_name = "PMCS預算編號記錄"
        wb.create_sheet(sheet_name, index=index)
        sheet = wb[sheet_name]
        for row in data1:
            sheet.append(row)
        index2 = 1
        sheet_name = "201领用"
        wb.create_sheet(sheet_name, index=index2)
        sheet = wb[sheet_name]
        for row2 in data2:
            sheet.append(row2)
        try:
            wb.save(os.path.join(settings.MEDIA_CHANGE_ROOT, filename))
            file_url = request.build_absolute_uri(settings.MEDIA_CHANGE_URL + filename)
            data = [file_url]
            return restful.ok(data=data)
        except:
            return restful.params_error(message=" please select one information ")
