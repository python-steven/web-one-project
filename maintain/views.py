from django.shortcuts import render
from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from datetime import datetime,timedelta
from django.conf import settings
import random
import string
import os
import pytz
import time
from openpyxl import load_workbook,Workbook
import json
#进入保养页面的获取数据以及设置保养次数和天数
class maintain_equipment_info(View):
    @csrf_exempt
    def get(self,request):
        #这里是获取近一周的数据，但是由于数据没有更新，所以捞取全部当前300天的数据，后面进行修改天数
        end_time = datetime.now()
        delta = timedelta(days=300)
        start_time = end_time-delta
        dict_data={}
        data = PartItem.objects.order_by("Id").filter(TrnDate__range=(start_time, end_time))
        data = list(data.values())
        # start_time = str(datetime.now()).split(' ')[0]
        # start_time = datetime.strptime(start_time, "%Y-%m-%d")
        # #计算保养次数和保养日期达不达标
        # for i in range(0,len(data)):
        #     if data[i]['NextCheckDate'] == None and data[i]['NextCheckCount'] ==0:
        #         data[i]['NextCheckDate'] = data[i]['CreatedTime'].split(' ')[0]
        #         data[i]['stand_date'] = 0
        #         data[i]['stand_count'] = 0
        #     else:
        #         time_u = datetime.strptime(str(data[i]['NextCheckDate']).split(' ')[0], "%Y-%m-%d")
        #         days = time_u - start_time
        #         data[i]['stand_date'] = days.days
        #         data[i]['stand_count'] = data[i]['NextCheckCount'] - data[i]['UsedTimes']
        dict_data['data'] = data
        try:
            return restful.ok(data=dict_data)
        except Exception as e:
            return restful.params_error(message=repr(e))

    #设置预警次数和预警天数
    @csrf_exempt
    def post(self,request):
        maintain_count = request.POST['maintain_count']
        maintain_date = request.POST['maintain_date']
        maintain_receiver = request.POST.getlist('maintain_receiver[]')
        maintain_receiver = list(maintain_receiver)
        try:
            parameter_count = Configuration.objects.filter(Type="mt_count")
            parameter_date = Configuration.objects.filter(Type="mt_date")
            mail_receiver_count = ",".join(maintain_receiver)
            if parameter_count and parameter_date:
                Configuration.objects.filter(Type="mt_count").update(Max=maintain_count,Min=0,Reminders=mail_receiver_count)
                Configuration.objects.filter(Type="mt_date").update(Max=maintain_date,Min=0,Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
            else:
                Configuration.objects.create(Type="mt_count", Max=maintain_count, Min=0, Reminders=mail_receiver_count)
                Configuration.objects.create(Type="mt_date", Max=maintain_date, Min=0, Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter create success")
        except:
            return restful.params_error(message="setup information error")

#单独的SN的保养更改
def maintain_setup_info(request):
    if request.method == "POST":
        main_count = request.POST['main_count']
        main_cycle = request.POST['main_cycle']
        main_date = request.POST['main_date']
        main_sn = request.POST['main_sn']
        try:
            PartItem.objects.filter(SN=main_sn).update(CheckCycleCount=main_count,CheckCycle=main_cycle,NextCheckDate=main_date)
            return restful.ok(message="maintain modify success")
        except Exception as e:
            return restful.params_error(message=repr(e))

#查询数据的函数
def maintain_query_part_name_data(request):
    if request.method == "POST":
        start_time = request.POST['main_start_time']
        end_t = request.POST['main_end_time']
        SN = str(request.POST['main_sn'])
        Part_name = str(request.POST['main_partname'])
        Status = request.POST['main_status']
        end_time = datetime.now()
        current_time = end_time.strftime('%Y-%m-%d')
        count = Configuration.objects.get(Type="mt_count")
        date = Configuration.objects.get(Type="mt_date")
        date = int(date.Max)
        delta = timedelta(date)
        check_time = end_time + delta
        count = str(int(count.Max))
        limit_value1 = list(Configuration.objects.filter(Type="mt_count").values("Max", "Id"))
        limit_value2 = list(Configuration.objects.filter(Type="mt_date").values("Max", "Id"))
        sql = 'SELECT "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","NextCheckCount" FROM "PartItem" where 1=1 '
        visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where 1=1 '
        dict_data = {}
        if start_time !="":
            sql = sql+' AND "TrnDate" >=\'%{0}%\''.format(start_time)
            visual_sql = visual_sql+' AND "TrnDate" >=\'%{0}%\''.format(start_time)
        if end_t !="":
            sql = sql+' AND "TrnDate" <=\'%{0}%\''.format(end_t)
            visual_sql = visual_sql+' AND "TrnDate" <=\'%{0}%\''.format(end_t)
        if SN !="":
            sql = sql+' AND "SN" =\'' + SN + '\''
            visual_sql = visual_sql+' AND "SN" =\'' + SN + '\''
        if Part_name !="":
            sql = sql+' AND "PartName" =\'' + Part_name + '\''
            visual_sql = visual_sql+' AND "PartName" =\'' + Part_name + '\''
        if Status == "正常":
            sql = sql + 'AND "NextCheckCount"-"UsedTimes" > ' + count + ' AND "NextCheckDate" >\'' + check_time.strftime("%Y-%m-%d") + '\''
            visual_normal = visual_sql + 'AND "NextCheckCount"-"UsedTimes" > ' + count + ' AND "NextCheckDate" >\'' + check_time.strftime("%Y-%m-%d") + '\'GROUP BY "PartName"'
            cur = connection.cursor()
            cur.execute(visual_normal)
            normal = cur.fetchall()
            dict_data['normal'] = normal
        if Status == "预警":
            sql = sql + 'AND ("NextCheckCount"-"UsedTimes" <= ' + count + ' AND "NextCheckCount"-"UsedTimes" >= 0'
            sql = sql + ' OR  "NextCheckDate">=\'' + current_time + '\' AND "NextCheckDate" < \'' + check_time.strftime("%Y-%m-%d") + '\')'
            visual_sql2 = visual_sql + 'AND ("NextCheckCount"-"UsedTimes" <= ' + count + ' AND "NextCheckCount"-"UsedTimes" >= 0'
            visual_waring = visual_sql2 + ' OR  "NextCheckDate">=\'' + current_time + '\' AND "NextCheckDate" < \'' + check_time.strftime("%Y-%m-%d") + '\') GROUP BY "PartName"'
            cur = connection.cursor()
            cur.execute(visual_waring)
            warning = cur.fetchall()
            dict_data['warning'] = warning
        if Status == "超标":
            sql = sql + 'AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \'' + current_time + '\')'
            visual_danger = visual_sql + 'AND ("NextCheckCount"-"UsedTimes" <0 OR "NextCheckDate" < \'' + current_time + '\') GROUP BY "PartName"'
            cur = connection.cursor()
            cur.execute(visual_danger)
            danger = cur.fetchall()
            dict_data['danger'] = danger
        cur = connection.cursor()
        cur.execute(sql)
        data = cur.fetchall()
        for i in range(len(data)):
            if data[i][6] == None and data[i][7] ==0:
                data[i] = list(data[i])
                data[i].extend([0,0])
            else:
                query_time = datetime.strptime(str(datetime.now()).split(' ')[0],"%Y-%m-%d")
                time_u = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")
                stand_count = data[i][7] - data[i][4]
                stand_date = (time_u-query_time).days
                data[i]=list(data[i])
                data[i].extend([stand_count,stand_date])
        dict_data['data'] = data
        dict_data['limit_value1'] = limit_value1
        dict_data['limit_value2'] = limit_value2
        #圆柱需要的数据
        return restful.ok(data=dict_data)

#生成报表的数据
def maintain_record(request):
    if request.method == "POST":
        maintain_id = request.POST.getlist('maintain_id[]')
        # sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where c.relname = '%s' and a.attrelid = c.oid and a.attnum>0''' % 'PartItem'
        # sql2 = 'SELECT "SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate" FROM' \
        #        ' "PartItem" INNER JOIN "Configuration" on "Configuration"."Type"= "mt_count","Configuration"."Type"="mt_date" where '
        # if len(maintain_id) == 1:
        #     maintain_id = maintain_id[0]
        #     sql3 = sql2 + '"PartItem"."Id"=' + maintain_id
        # else:
        #     maintain_id = tuple(maintain_id)
        #     maintain_id = str(maintain_id)
        #     sql3 = sql2 + '"PartItem"."Id" in ' + maintain_id
        #
        # cur = connection.cursor()
        # cur.execute(sql3)
        # data = cur.fetchall()
        # cur.close()
        #
        # cur2 = connection.cursor()
        # cur2.execute(sql1)
        # statement_data = cur2.fetchall()
        # cur2.close()
        # statement_data = [attr[0] for attr in statement_data]
        #
        # data1 = []
        # data1.append(statement_data[statement_data.index("SN")])
        # data1.append(statement_data[statement_data.index("PartName")])
        # data1.append('Min')
        # data1.append('Max')
        # data1.append(statement_data[statement_data.index("NGRate")])
        # data1.append(statement_data[statement_data.index("ErrorCounts")])
        # data1.append(statement_data[statement_data.index("UsedTimes")])
        # data1.append("status")
        #
        # data.insert(0, data1)
        try:
            # sheet_name = "NG率监控表单"
            # time_num = int(time.time())
            # time_num = str(time_num)
            # file_root = settings.MEDIA_MONITOR_ROOT
            # file_url = settings.MEDIA_MONITOR_URL
            # filename = 'download' + time_num + '.xlsx'
            # data = statement_excle(request, data, sheet_name, file_root, file_url, filename)
            return restful.ok(data=maintain_id)
        except:
            return restful.params_error(message="download fail")
#生成报表并存入服务器的函数
def statement_excle(request,data,sheet_name,file_root,file_url,filename):
    # 写入文件
    # time_num = int(time.time())
    # time_num = str(time_num)
    # sheet_name = "预算表单1"
    # filename = 'download' + time_num + '.xlsx'

    wb = Workbook()
    index = 0
    wb.create_sheet(sheet_name, index=index)
    sheet = wb[sheet_name]
    for row in data:
        row = list(row)
        if len(row)==7:
            #正常
            if row[2] > row[4]:
                row.append("正常")
            #预警
            if row[2] <= row[4] <=row[3]:
                row.append("预警")
            #超标
            if row[3] < row[4]:
                row.append("超标")
        sheet.append(row)
    wb.save(os.path.join(file_root, filename))

    file_url = request.build_absolute_uri(file_url + filename)
    data = [file_url]
    return data

#by PN的批量更改
def maintain_setup_by_pn(request):
    if request.method == "POST":
        main_pn = request.POST['main_partname']
        main_count = request.POST['main_count']
        main_day = request.POST['main_day']
        main_date = request.POST['main_date']
        try:
            result = PartItem.objects.filter(PN=main_pn)
            if len(restful) >0:
                result.update(CheckCycleCount=main_count,CheckCycle=main_day,NextCheckDate=main_date)
                return restful.ok(message="maintain modify success")
            return restful.params_error(message="PN query is null")
        except:
            return restful.params_error(message="PN query is null")

#设备保养得监控中的按条件的捞取需要设置的items函数
def maintain_query_operation(request):
    if request.method =="POST":
        SN = str(request.POST['sn'])
        PN = str(request.POST['pn'])
        Status = request.POST['status']
        Next_maintain_time = request.POST['next_time']
        Next_maintain_time_1 = request.POST['next_time_1']
        try:
            dict_data={}
            end_time = datetime.now()
            current_time = end_time.strftime('%Y-%m-%d')
            count = Configuration.objects.get(Type="mt_count")
            date = Configuration.objects.get(Type="mt_date")
            date = int(date.Max)
            delta = timedelta(date)
            check_time = end_time + delta
            count = str(int(count.Max))
            limit_value1 = list(Configuration.objects.filter(Type="mt_count").values("Max", "Id"))
            limit_value2 = list(Configuration.objects.filter(Type="mt_date").values("Max", "Id"))
            sql = 'SELECT "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","NextCheckCount","PN" FROM "PartItem" where 1=1 '
            if SN !="":
                sql = sql+' AND "SN" =\'' + SN + '\''
            if PN !="":
                sql = sql+' AND "PN" =\'' + PN + '\''
            if Status == "正常":
                sql = sql + 'AND "NextCheckCount"-"UsedTimes" > '+count+' AND "NextCheckDate" >\''+check_time.strftime("%Y-%m-%d")+'\''
            if Status == "预警":
                sql = sql + 'AND ("NextCheckCount"-"UsedTimes" <= ' + count + ' AND "NextCheckCount"-"UsedTimes" >= 0'
                sql = sql + ' OR  "NextCheckDate">=\'' + current_time + '\' AND "NextCheckDate" < \'' + check_time.strftime("%Y-%m-%d") + '\')'
            if Status == "超期":
                sql = sql + 'AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \''+current_time+'\')'
            if Next_maintain_time !="":
                sql = sql+' AND "NextCheckDate" >= \''+Next_maintain_time+'\' AND "NextCheckDate" <= \''+Next_maintain_time_1+'\''
            cur = connection.cursor()
            cur.execute(sql)
            data = cur.fetchall()
            for i in range(len(data)):
                if data[i][6] == None and data[i][7] == 0:
                    data[i] = list(data[i])
                    data[i].extend([0, 0])
                else:
                    query_time = datetime.strptime(str(datetime.now()).split(' ')[0], "%Y-%m-%d")
                    time_u = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")
                    stand_count = data[i][7] - data[i][4]
                    stand_date = (time_u - query_time).days
                    data[i] = list(data[i])
                    data[i].extend([stand_count, stand_date])
            dict_data['limit_value1'] = limit_value1
            dict_data['limit_value2'] = limit_value2
            dict_data['data'] = data
            return restful.ok(data=dict_data)
        except:
            return restful.params_error(message='need setup maintain data')

#用户使用保养得页面进行保养动作的函数
def maintain_query_maintain(request):
    if request.method == "POST":
        maintain_ids = request.POST.getlist('statement_mt[]')
        maintain_date = request.POST['maintain_date']
        maintain_operator = request.POST['maintain_operator']
        maintain_status = request.POST['maintain_status']
        maintain_text = request.POST['maintain_text']
        data ={
            'maintain_ids':maintain_ids,
            'maintain_date':maintain_date,
            'maintain_operator':maintain_operator,
            'maintain_status':maintain_status,
            'maintain_text':maintain_text,
        }
        try:
            return restful.ok(data=data)
        except:
            return restful.params_error(message="maintain is null")

# def monitor_query_info(request):
#     if request.method == "POST":
#         sn = str(request.POST['sn'])
#         part_name = request.POST['part_name']
#         status = request.POST['status']
#         start_time = request.POST['start_tim']
#         end_time = request.POST['end_tim']
#         NG_monitor_type = "NG率设定"
#         sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
#               ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
#               'FROM "PartItem" WHERE 1=1'
#         visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where 1=1 '
#         try:
#             parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
#             limit_value = [parmeter_stands.Min, parmeter_stands.Max]
#             parmeter_stands_min = str(parmeter_stands.Min)
#             parmeter_stands_max = str(parmeter_stands.Max)
#             dict_data = {}
#             normal = []
#             warning = []
#             danger = []
#             if status == "":
#                 if start_time != "":
#                     sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
#                 if end_time != "":
#                     sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
#                 if part_name != "":
#                     sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
#                     visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
#                 if sn != "":
#                     sql = sql + 'AND "SN" = \'' + sn + '\''
#                     visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
#                 # 正常
#                 visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
#                 # 预警
#                 visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
#                                     + 'AND "NGRate" <= \'' + parmeter_stands_max + '\' group by "PartName"'
#                 # 超标
#                 visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
#                 cur = connection.cursor()
#                 cur.execute(visual_sql_normal)
#                 normal = cur.fetchall()
#
#                 cur = connection.cursor()
#                 cur.execute(visual_sql_waring)
#                 warning = cur.fetchall()
#
#                 cur = connection.cursor()
#                 cur.execute(visual_sql_danger)
#                 danger = cur.fetchall()
#
#                 cur = connection.cursor()
#                 cur.execute(sql)
#                 data = cur.fetchall()
#
#                 dict_data['data'] = data
#                 dict_data['limit_value'] = limit_value
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 return restful.ok(data=dict_data)
#             else:
#                 if sn != "":
#                     sql = sql + 'AND "SN" = \'' + sn + '\''
#                     visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
#                 if part_name != "":
#                     sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
#                     visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
#                 if start_time != "":
#                     sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
#                 if end_time != "":
#                     sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
#                 if status == "正常":
#                     sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
#                     visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_normal)
#                     normal = cur.fetchall()
#
#                 if status == "预警":
#                     sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' + parmeter_stands_max + '\''
#                     visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' + parmeter_stands_max + '\'group by "PartName"'
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_waring)
#                     warning = cur.fetchall()
#
#                 if status == "超标":
#                     sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
#                     visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_danger)
#                     danger = cur.fetchall()
#
#                 cur = connection.cursor()
#                 cur.execute(sql)
#                 data = cur.fetchall()
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 dict_data['data'] = data
#                 dict_data['limit_value'] = limit_value
#                 return restful.ok(data=dict_data)
#         except:
#             return restful.params_error(message="query data is empty")